from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string

wb = load_workbook("C:/Users/uilov/Desktop/개발 관련 폴더/추가모집관련/ex.xlsx", data_only=True)
ws = wb.worksheets[0]
wb.create_sheet(title="정리본")
wss = wb["정리본"]
num = 1

for a in range(ws.max_row - 6) :
    # 한줄로 DB화 코드(1) : 모집인원 미지정 시(전체 0명)
    if int(ws.cell(row = a + 7, column = 14).value) == 0 :
        if ws.cell(row = a + 7, column = 9).value :
            # 한줄로 DB화 코드(1-2) : 인원 지정 시(0명 차례대로 깔기위함) # 모집단위 0명을 깔기 위함
            joen = ws.cell(row=a + 7, column=9).value
            Hackwa = []
            c = 0

            for b in range(len(ws.cell(row=a + 7, column=9).value)) :
                if joen[b] == ',' :
                    try :
                        if int(joen[b-1]) > -1 :
                            Hackwa.append(joen[c:b])
                            c = b + 1
                    except : 
                        if joen[b-1].isdigit == True :
                            Hackwa.append(joen[c:b])
                            c = b + 1

                if b == len(ws.cell(row=a + 7, column=9).value) -2 :
                    Hackwa.append(joen[c:len(ws.cell(row=a + 7, column=9).value)])

            for i in Hackwa :
                for d in range(8) :
                    wss.cell(row = num, column = d + 1, value = ws.cell(row=a + 7, column= d + 1).value)

                wss.cell(row = num, column = 9, value = i)
                num = num + 1

        #숫자 찢어서 넣기                
            for i in range(num-len(Hackwa), num) :
                joen2 = wss.cell(row = i, column = 9).value
                for ii in range(len(str(wss.cell(row = i, column = 9).value))) :
                    if joen2[ii] == ':' :
                        wss.cell(row = i, column = 9, value=joen2[0:ii])
                        wss.cell(row = i, column = 10, value=joen2[ii+1:len(joen2)])

        else :
            # 한줄로 DB화 코드(1-1) : 모집인원 미지정 시(모집단위 없음))
            for d in range(8) :
                wss.cell(row = num, column = d + 1, value = ws.cell(row=a + 7, column= d + 1).value)
            
            wss.cell(row = num, column = 9, value ="미지정")
            wss.cell(row = num, column = 10, value =0)
            num = num + 1

    elif ws.cell(row = a + 7, column = 12).value or ws.cell(row = a + 7, column = 13).value :
        # 한줄로 DB화 코드(3) : 미지정 모집
        joen = ws.cell(row=a + 7, column=9).value
        Hackwa = []
        c = 0
        Hackwa_excep = [] 

        for b in range(len(ws.cell(row=a + 7, column=9).value)) :
            if joen[b] == ',' :
                try :
                    if int(joen[b-1]) > -1 :
                        Hackwa.append(joen[c:b])
                        c = b + 1
                except : 
                    if joen[b-1].isdigit == True :
                        Hackwa.append(joen[c:b])
                        c = b + 1

            if b == len(ws.cell(row=a + 7, column=9).value) -2 :
                Hackwa.append(joen[c:len(ws.cell(row=a + 7, column=9).value)])

        for i in Hackwa :
            # 몇몇에만 인원 수 배정한 학교
            if i[str(i).find(':')+1:len(i)] == "0" or i[str(i).find(':')+1:len(i)] == 0 or i[str(i).find(':')+1:len(i)] == " 0" or i[str(i).find(':')+1:len(i)] == "0 " :
                Hackwa_excep.append(i)
            else :
                for d in range(8) :
                    wss.cell(row = num, column = d + 1, value = ws.cell(row=a + 7, column= d + 1).value)

                wss.cell(row = num, column = 9, value = i)
        
            #숫자 찢어서 넣기                
                joen2 = wss.cell(row = num, column = 9).value       
                for ii in range(len(str(wss.cell(row = num, column = 9).value))) :
                    if joen2[ii] == ':' :
                        wss.cell(row = num, column = 9, value=joen2[0:ii])
                        wss.cell(row = num, column = 10, value=joen2[ii+1:len(joen2)])

                num = num + 1


        if Hackwa_excep == [] :
            print("예외학과 비었음")
        else :
            Hackwa_excep2 = [] 
            for d in range(8) :
                wss.cell(row = num, column = d + 1, value = ws.cell(row=a + 7, column= d + 1).value)

            for i in Hackwa_excep :
                joen2 = i
                for ii in range(len(i)) :
                    if joen2[ii] == ':' :
                       Hackwa_excep2.append(joen2[0:ii])

            Hackwa_excepall = ", ".join(Hackwa_excep2)

            wss.cell(row = num, column=9, value =Hackwa_excepall)
            if ws.cell(row = a + 7, column = 12).value :
                if ws.cell(row = a + 7, column = 13).value :
                    Tong = int(ws.cell(row = a + 7, column = 12).value) + int(ws.cell(row = a + 7, column = 13).value)
                else :
                    Tong = ws.cell(row = a + 7, column = 12).value
            else :
                Tong = ws.cell(row = a + 7, column = 13).value
            wss.cell(row=num, column=10, value = Tong)
            
            num = num + 1


            Hackwa_excep = []
    else :
    # 한줄로 DB화 코드(2) : 전 인원 지정 시
        joen = ws.cell(row=a + 7, column=9).value
        Hackwa = []
        c = 0

        for b in range(len(ws.cell(row=a + 7, column=9).value)) :
            if joen[b] == ',' :
                try :
                    if int(joen[b-1]) > -1 :
                        Hackwa.append(joen[c:b])
                        c = b + 1
                except : 
                    if joen[b-1].isdigit == True :
                        Hackwa.append(joen[c:b])
                        c = b + 1

            if b == len(ws.cell(row=a + 7, column=9).value) -2 :
                Hackwa.append(joen[c:len(ws.cell(row=a + 7, column=9).value)])

        for i in Hackwa :
            for d in range(8) :
                wss.cell(row = num, column = d + 1, value = ws.cell(row=a + 7, column= d + 1).value)

            wss.cell(row = num, column = 9, value = i)
            num = num + 1

        for i in range(num-len(Hackwa), num) :
            joen2 = str(wss.cell(row = i, column = 9).value)
            for ii in range(len(str(wss.cell(row = i, column = 9).value))) :
                if joen2[ii] == ':' :
                    wss.cell(row = i, column = 9, value = joen2[0:ii])
                    wss.cell(row = i, column = 10, value = joen2[ii+1:len(joen2)])


wb.save("extest.xlsx")