import os
import PyPDF2
from openpyxl import workbook
from openpyxl import load_workbook
 

# PDF파일만 추출
path = "C:/Users/{}/desktop".format(os.getlogin()) # {}부분에 사용자 이름
All_file_list = os.listdir(path + "/PDF암호화")
file_list = []
people = []
password = []

for a in All_file_list :
    if a[len(a) - 4:len(a)] == ".pdf" :
        file_list.append(a)

file_list_number = len(file_list)

wb = load_workbook(path + "/PDF암호화/PDF_암호.xlsx", data_only=True)
ws = wb.worksheets[0]

for b in range(ws.max_row) :
    people.append(ws.cell(row = b+1, column = 1).value)
    password.append(ws.cell(row = b+1, column = 2).value)

for d in range(len(password)) :
    password[d] = password[d][1:len(password[d])-1]

for i in range(file_list_number) :
    Original_file = path + "/PDF암호화/" + file_list[i]
    New_file = path + "/PDF암호화/(R)" + file_list[i]


    # 원본 파일 불러오기 (첫번째 파일)
    pdfFile = open(Original_file, 'rb')
    pdfReader = PyPDF2.PdfFileReader(pdfFile)

    Encrypted = pdfReader.isEncrypted  # 암호가 걸려 있으면 - True // 걸려있지 않으면 - False

    if Encrypted :
        print()
    else :
        for c in range(len(people)) :
            if file_list[i][12:15] == people[c] :
                key = c
                break

        try :
            R_password = password[key]
        except :
            R_password = "111111"

        # 값을 붙여넣을 빈 객체(파일)를 생성
        pdfWriter = PyPDF2.PdfFileWriter()

        # 원본 pdf의 각 page의 값을 하나씩 빈 객체에 붙여넣기
        for pageNum in range(pdfReader.numPages):
            pdfWriter.addPage(pdfReader.getPage(pageNum))

        # 빈 객체를 암호화 시킴
        pdfWriter.encrypt(R_password)

        # 빈 객체에 붙여넣은 값들을 새로운 pdf 파일에 저장하기
        pdfFile_new = open(New_file, 'wb')
        pdfWriter.write(pdfFile_new)

        # 원본 파일 및 새로 생성한 pdf 파일을 닫기
        pdfFile_new.close()
        pdfFile.close()

        os.remove(path + "/PDF암호화/" + file_list[i])

        key = ""


print("완료하였습니다")
os.system("pause")